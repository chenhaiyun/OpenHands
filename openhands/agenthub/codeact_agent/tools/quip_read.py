import os
import csv
import io
import re
import requests
import pandas as pd
from bs4 import BeautifulSoup
from typing import Optional, List, Dict, Any, Union, Tuple
from openpyxl import load_workbook
from litellm import ChatCompletionToolParam, ChatCompletionToolParamFunctionChunk
from openhands.core.logger import openhands_logger as logger

_QUIP_DESCRIPTION = """Read content from a Quip document, especially spreadsheets. This tool allows you to read Quip documents using their thread IDs.

You can use this tool to:
1. Read Quip spreadsheet content
2. Get spreadsheet metadata (rows, columns)
3. Access specific sheets within a spreadsheet

You can provide the thread ID in these formats:
- Direct thread ID: ABCDEFGHIJKLM
- Quip URL format: quip://ABCDEFGHIJKLM?sheet=SheetName
"""

class QuipClient:
    """
    Simple Quip API client implementation
    """
    def __init__(self, access_token: str, base_url: str = "https://platform.quip.com"):
        """
        Initialize the Quip client with the given access token and base URL
        """
        self.access_token = access_token
        self.base_url = base_url.rstrip('/')
        self.session = requests.Session()
        self.session.headers = {
            'Authorization': f'Bearer {access_token}'
        }
        logger.info(f"QuipClient initialized with base URL: {self.base_url}")

    def get_thread(self, thread_id: str) -> Dict[str, Any]:
        """Get a thread by ID"""
        logger.info(f"Getting thread: {thread_id}")
        response = self.session.get(f"{self.base_url}/1/threads/{thread_id}")
        response.raise_for_status()
        return response.json()

    def export_thread_to_xlsx(self, thread_id: str, output_path: str) -> str:
        """Export a thread to XLSX format"""
        logger.info(f"Exporting thread {thread_id} to XLSX")
        response = self.session.get(
            f"{self.base_url}/1/threads/{thread_id}/export/xlsx",
            stream=True
        )
        response.raise_for_status()
        
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        with open(output_path, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
        
        logger.info(f"Successfully exported XLSX to {output_path}")
        return output_path

    def export_thread_to_csv_fallback(self, thread_id: str, sheet_name: Optional[str] = None) -> str:
        """Export a thread to CSV using HTML parsing as fallback"""
        logger.info(f"Using fallback method to export thread {thread_id} to CSV")
        
        thread = self.get_thread(thread_id)
        if not thread or 'html' not in thread:
            raise ValueError("Could not retrieve thread or thread has no HTML content")

        sheet = find_sheet_by_name(thread['html'], sheet_name)
        if not sheet:
            if sheet_name:
                raise ValueError(f"Could not find sheet '{sheet_name}' in the document")
            else:
                raise ValueError("Could not find any spreadsheet in the document")

        data = extract_sheet_data(sheet)
        if not data:
            raise ValueError(f"No data found in sheet '{sheet_name or 'default'}'")

        csv_buffer = io.StringIO()
        writer = csv.writer(csv_buffer, quoting=csv.QUOTE_MINIMAL)
        for row in data:
            writer.writerow(row)
        
        csv_data = csv_buffer.getvalue()
        csv_buffer.close()
        return csv_data

    def is_spreadsheet(self, thread_id: str) -> bool:
        """Check if a thread is a spreadsheet"""
        try:
            thread = self.get_thread(thread_id)
            if not thread or 'thread' not in thread:
                return False
            thread_type = thread.get('thread', {}).get('type', '').lower()
            return thread_type == 'spreadsheet'
        except Exception as e:
            logger.error(f"Error checking if thread is spreadsheet: {str(e)}")
            return False

def find_sheet_by_name(document_html: str, sheet_name: Optional[str] = None) -> Optional[Any]:
    """Find a spreadsheet with the given name in the document HTML"""
    soup = BeautifulSoup(document_html, 'html.parser')
    
    if sheet_name:
        table = soup.find('table', attrs={'title': sheet_name})
        if table:
            return table
        
        for heading in soup.find_all(['h1', 'h2', 'h3']):
            if heading.get_text().strip() == sheet_name:
                next_table = heading.find_next('table')
                if next_table:
                    return next_table
    
    return soup.find('table')

def is_metadata_row(row: List[str]) -> bool:
    """Determine if a row is likely metadata"""
    non_empty_cells = sum(1 for cell in row if cell.strip())
    if non_empty_cells <= 1:
        return True
        
    date_indicators = ["updated on", "created on", "modified on", "as of"]
    row_text = " ".join(row).lower()
    return any(indicator in row_text for indicator in date_indicators)

def is_header_row(row: List[str]) -> bool:
    """Determine if a row is likely a header row"""
    if not row or not any(cell.strip() for cell in row):
        return False
        
    max_header_length = 50
    sentence_punctuation = ['.', '!', '?']
    list_indicators = ['•', '-', '1)', 'a)', '1.', 'i.', 'i)']
    
    for cell in row:
        cell = cell.strip()
        if not cell:
            continue
            
        if len(cell) > max_header_length:
            return False
            
        if any(p in cell[1:-1] for p in sentence_punctuation):
            return False
            
        if any(cell.lower().startswith(indicator) for indicator in list_indicators):
            return False
    
    return True

def extract_sheet_data(sheet: Any) -> List[List[str]]:
    """Extract data from a sheet element"""
    if sheet is None:
        return []
    
    rows_data = []
    for tr in sheet.find_all('tr'):
        cols = tr.find_all('td')
        if not cols:
            continue
            
        row_text = [col.get_text().strip() for col in cols]
        if not any(text for text in row_text):
            continue
            
        cleaned_row = []
        for text in row_text:
            if text and not text.isdigit():
                text = text.replace('\u200b', '').strip()
                cleaned_row.append(text)
        
        if cleaned_row:
            rows_data.append(cleaned_row)
    
    if not rows_data:
        return []
        
    start_idx = 0
    while start_idx < len(rows_data) and is_metadata_row(rows_data[start_idx]):
        start_idx += 1
        
    if start_idx >= len(rows_data):
        return []
        
    header_idx = start_idx
    while header_idx < len(rows_data):
        if is_header_row(rows_data[header_idx]):
            header_row = rows_data[header_idx]
            feature_col_idx = header_row.index('Feature to Address') if 'Feature to Address' in header_row else None
            break
        header_idx += 1
    else:
        header_idx = start_idx - 1
        feature_col_idx = None
    
    rows = []
    for i in range(start_idx):
        rows.append(rows_data[i])
    
    if header_idx >= start_idx:
        rows.append(rows_data[header_idx])
    
    for row_data in rows_data[header_idx + 1:]:
        if not is_metadata_row(row_data):
            processed_row = []
            for col_idx, cell in enumerate(row_data):
                cell_text = cell.strip()
                
                if feature_col_idx is not None and col_idx == feature_col_idx:
                    if cell_text.startswith('a)'):
                        for i in range(97, 122):
                            char = chr(i)
                            next_char = chr(i+1)
                            cell_text = cell_text.replace(f"{char}){next_char})", f"{char})\n{next_char})")
                
                processed_row.append(cell_text)
            
            if any(cell.strip() for cell in processed_row):
                rows.append(processed_row)
    
    return rows

def convert_xlsx_to_csv(xlsx_path: str, sheet_name: Optional[str] = None) -> str:
    """Convert XLSX file to CSV format"""
    logger.info(f"Reading XLSX file from {xlsx_path}")
    
    wb = load_workbook(filename=xlsx_path, read_only=False, data_only=True)
    sheet_names = wb.sheetnames
    logger.info(f"Available sheets: {', '.join(sheet_names)}")
    
    target_sheet = None
    if sheet_name:
        if sheet_name in sheet_names:
            target_sheet = wb[sheet_name]
        else:
            sheet_lower = sheet_name.lower()
            for s in sheet_names:
                if s.lower() == sheet_lower:
                    target_sheet = wb[s]
                    break
            
            if not target_sheet:
                raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(sheet_names)}")
    else:
        target_sheet = wb.active
    
    csv_buffer = io.StringIO()
    csv_writer = csv.writer(csv_buffer)
    
    from openpyxl.utils import get_column_letter
    max_col = 0
    for col in target_sheet.columns:
        max_col = max(max_col, col[0].column)
    
    logger.info(f"Found {max_col} columns")
    column_letters = [get_column_letter(i) for i in range(1, max_col + 1)]
    
    for row_idx, row in enumerate(target_sheet.rows, 1):
        row_data = []
        for col in column_letters:
            cell = target_sheet[f"{col}{row_idx}"]
            value = cell.value
            if isinstance(value, str):
                value = value.strip()
            row_data.append('' if value is None else str(value))
            
        if row_idx <= 5:
            logger.info(f"Row {row_idx} data: {row_data}")
            
        csv_writer.writerow(row_data)
    
    csv_data = csv_buffer.getvalue()
    csv_buffer.close()
    wb.close()
    
    return csv_data

def _parse_quip_url(url: str) -> Tuple[str, Optional[str]]:
    """
    Parse a Quip URL to extract thread ID and sheet name.
    
    Args:
        url (str): The Quip URL or thread ID (e.g., "quip://ABCDEFGHIJKLM?sheet=SheetName")
        
    Returns:
        Tuple[str, Optional[str]]: Thread ID and optional sheet name
    """
    # Check if it's a Quip URL format
    if url.startswith('quip://'):
        # Extract thread ID and optional sheet name
        match = re.match(r'quip://([^?]+)(?:\?sheet=(.+))?', url)
        if match:
            thread_id = match.group(1)
            sheet_name = match.group(2) if match.group(2) else None
            logger.info(f"Parsed Quip URL - Thread ID: {thread_id}, Sheet: {sheet_name}")
            return thread_id, sheet_name
    
    # If not a Quip URL, assume it's a direct thread ID
    logger.info(f"Using direct thread ID: {url}")
    return url, None

def _read_quip_spreadsheet(thread_id: str, sheet_name: Optional[str] = None) -> dict:
    """Read content from a Quip spreadsheet"""
    # Parse thread_id if it's in URL format
    parsed_thread_id, parsed_sheet_name = _parse_quip_url(thread_id)
    # Use parsed sheet name if provided in URL and not overridden by parameter
    sheet_name = sheet_name or parsed_sheet_name
    
    logger.info(f"Reading Quip spreadsheet - Thread ID: {parsed_thread_id}, Sheet: {sheet_name}")
    
    # Get Quip API token from environment
    api_token = os.getenv("QUIP_TOKEN")
    if not api_token:
        raise ValueError("QUIP_TOKEN environment variable is not set")

    # Get Quip base URL from environment or use default
    base_url = os.getenv("QUIP_BASE_URL", "https://platform.quip.com")
    
    # Initialize client
    client = QuipClient(api_token, base_url)
    
    try:
        # Check if it's a spreadsheet
        if not client.is_spreadsheet(parsed_thread_id):
            raise ValueError(f"Thread {parsed_thread_id} is not a spreadsheet")
            
        # Get workspace directory from config or use default
        workspace_dir = os.getenv("OPENHANDS_WORKSPACE", "./workspace")
        os.makedirs(workspace_dir, exist_ok=True)
        
        # Try XLSX export first
        temp_xlsx = os.path.join(workspace_dir, f"{parsed_thread_id}.xlsx")
        try:
            xlsx_path = client.export_thread_to_xlsx(parsed_thread_id, temp_xlsx)
            csv_content = convert_xlsx_to_csv(xlsx_path, sheet_name)
            os.remove(xlsx_path)  # Clean up temp file
        except Exception as e:
            logger.warning(f"XLSX export failed, falling back to HTML parsing: {str(e)}")
            csv_content = client.export_thread_to_csv_fallback(parsed_thread_id, sheet_name)
        
        # Get thread data for metadata
        thread_data = client.get_thread(parsed_thread_id)
        
        # Convert CSV to DataFrame for metadata
        df = pd.read_csv(io.StringIO(csv_content))
        
        metadata = {
            "rows": len(df),
            "columns": len(df.columns),
            "sheet_name": sheet_name or "default",
            "title": thread_data.get("thread", {}).get("title", "")
        }
        
        logger.info(f"Successfully read spreadsheet - Metadata: {metadata}")
        return {
            "csv_content": csv_content,
            "metadata": metadata
        }

    except Exception as e:
        raise Exception(f"Error processing Quip spreadsheet: {str(e)}")

QuipReadTool = ChatCompletionToolParam(
    type='function',
    function=ChatCompletionToolParamFunctionChunk(
        name='quip_read',
        description=_QUIP_DESCRIPTION,
        parameters={
            'type': 'object',
            'properties': {
                'thread_id': {
                    'type': 'string',
                    'description': 'The Quip document thread ID or URL (e.g., "ABCDEFGHIJKLM" or "quip://ABCDEFGHIJKLM?sheet=SheetName").',
                },
                'sheet_name': {
                    'type': 'string',
                    'description': 'Optional name of the specific sheet to read from the spreadsheet. If provided in the URL, this parameter will override it.',
                }
            },
            'required': ['thread_id'],
        },
    ),
)
